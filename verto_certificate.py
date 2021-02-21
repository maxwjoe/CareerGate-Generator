# *** IMPORT MODULES ***
from datetime import datetime
import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import  TA_CENTER
from reportlab.lib import colors

# *** GLOBALS ***
excel_file = "apprentice_data.xlsx"
column_count = 5
start_row = 4
apprentice_count = 0
error_counter = 0

# *** SETUP ***
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
pdfmetrics.registerFont(TTFont('ArialBold', 'arialbd.ttf'))
pdfmetrics.registerFont(TTFont('TNRB', 'timesbd.ttf'))
pdfmetrics.registerFont(TTFont('ARIALROUNDEDBOLD', 'ARLRDBD.ttf'))

# *** FUNCTION DEFINITIONS ***
def check_for_worksheet():
    exist_flag = 0
    if (os.path.isfile(excel_file)):
        exist_flag = 1

    return exist_flag


def generate_new_data_sheet():
    wb = Workbook()
    sheet = wb.active

    # SETUP COLUMNS
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20

    # ADD TITLE
    sheet.cell(1,1).value = "APPRENTICE INFORMATION"
    sheet.cell(1,1).font = Font(bold = True, size = 30, color = 'FF0000')

    # ADD BORDER FILL
    redFill = PatternFill(start_color='f8fc00',
                          end_color='f8fc00',
                          fill_type='solid')

    blackFill = PatternFill(start_color='000000',
                            end_color='000000',
                            fill_type='solid')


    # ADD COLUMN HEADINGS
    sheet.cell(3,1).value = "FIRST NAME"
    sheet.cell(3,2).value = "LAST NAME"
    sheet.cell(3,3).value = "SCORE"
    sheet.cell(3,4).value = "DATE TAKEN"
    sheet.cell(3,5).value = "TIME ELAPSED"

    # SET BOLD FONT AND ADD BORDER
    for i in range(1,6):
        sheet.cell(3,i).font = Font(bold = True, size=16)
        sheet.cell(2,i).fill = redFill
        sheet.cell(1,i).fill = blackFill

    wb.save(filename=excel_file)

def get_apprentice_count(sheet):
    count = 0
    while sheet.cell(start_row + count, 1).value != None:
        count += 1
    return count


def get_data(sheet, row):
    global error_counter
    ERROR_FLAG = 0
    array = [0] * column_count
    for i in range(0, column_count):
        array[i] = sheet.cell(row, i + 1).value
        if array[i] == None and ERROR_FLAG != 1:
            error_row_array[error_counter] = row
            error_counter += 1
            ERROR_FLAG = 1
            if i == 0:
                array[i] = "FIRST NAME"
            elif i == 1:
                array[i] = "LAST NAME"
            elif i== 2:
                array[i] = -1
            elif i == 3:
                now = datetime.now()
                current_time = now.strftime("%d/%m/%y")
                array[i] = str(current_time)
            elif i == 4:
                array[i] = "5"

    return array


def generate_pdf(apprentice_vector):

    #Inital PDF Setup
    pdf = SimpleDocTemplate(str(apprentice_vector[0]) + " " + str(apprentice_vector[1]) + " Certificate.pdf",
                            pagesize = A4, rightMargin = 72, leftMargin=72,
                            topMargin = 72, bottomMargin = 18)
    stylesheet = getSampleStyleSheet()
    story = []

    # STYLES
    stylesheet.add(ParagraphStyle(name='Heading_CENTER',
                                  parent=stylesheet['Normal'],
                                  fontName='ARIALROUNDEDBOLD',
                                  wordWrap='LTR',
                                  alignment=TA_CENTER,
                                  fontSize=22,
                                  leading=13,
                                  textColor=colors.dimgray,
                                  borderPadding=0,
                                  leftIndent=0,
                                  rightIndent=0,
                                  spaceAfter=0,
                                  spaceBefore=0,
                                  splitLongWords=True,
                                  spaceShrinkage=0.05,
                                  ))
    stylesheet.add(ParagraphStyle(name='Text_CENTER',
                                  parent=stylesheet['BodyText'],
                                  fontName='Arial',
                                  wordWrap='',
                                  alignment=TA_CENTER,
                                  fontSize=11,
                                  leading=13,
                                  textColor=colors.black,
                                  borderPadding=0,
                                  leftIndent=-45,
                                  rightIndent=-45,
                                  spaceAfter=0,
                                  spaceBefore=0,
                                  splitLongWords=True,
                                  spaceShrinkage=0.05,
                                  ))

    stylesheet.add(ParagraphStyle(name='Text_CENTER_BOLD',
                                  parent=stylesheet['BodyText'],
                                  fontName='ArialBold',
                                  wordWrap='',
                                  alignment=TA_CENTER,
                                  fontSize=11,
                                  leading=13,
                                  textColor=colors.black,
                                  borderPadding=0,
                                  leftIndent=-45,
                                  rightIndent=-45,
                                  spaceAfter=0,
                                  spaceBefore=0,
                                  splitLongWords=True,
                                  spaceShrinkage=0.05,
                                  ))

    stylesheet.add(ParagraphStyle(name='Result_Heading_CENTER',
                                  parent=stylesheet['BodyText'],
                                  fontName='Arial',
                                  wordWrap='',
                                  alignment=TA_CENTER,
                                  fontSize=11,
                                  leading=13,
                                  textColor=colors.black,
                                  borderPadding=0,
                                  leftIndent=-45,
                                  rightIndent=-45,
                                  spaceAfter=0,
                                  spaceBefore=0,
                                  splitLongWords=True,
                                  spaceShrinkage=0.05,
                                  ))
    stylesheet.add(ParagraphStyle(name='Name_CENTER',
                                  parent=stylesheet['BodyText'],
                                  fontName='TNRB',
                                  wordWrap='',
                                  alignment=TA_CENTER,
                                  fontSize=36,
                                  leading=13,
                                  textColor=colors.black,
                                  borderPadding=0,
                                  leftIndent=-45,
                                  rightIndent=-45,
                                  spaceAfter=0,
                                  spaceBefore=0,
                                  splitLongWords=True,
                                  spaceShrinkage=0.05,
                                  ))

    #COMMENT LOGIC
    comment = " "
    int_score = int(apprentice_vector[2])
    if int_score < 50:
        comment = "Moderate-Low Knowledge"
    elif int_score >= 50 and int_score < 65:
        comment = "Moderate Knowledge"
    elif int_score >= 65 and int_score < 75:
        comment = "Good Knowledge"
    elif int_score >= 75 and int_score < 85:
        comment = "Strong Knowledge"
    elif int_score >= 85:
        comment = "Very Strong Knowledge"

    # ADD LOGO
    logo = "Files\CareerGate_Logo.JPG"
    im = Image(logo, 14 * cm, 3 * cm)
    im.__setattr__("_offs_x", 0)
    im.__setattr__("_offs_y", 30)
    story.append(im)

    # TEXT
    h1 = 'Certificate of Participation'
    p1 = "Thank you for registering your interest in an apprenticeship/traineeship on CareerGate. Your initiative to pursue further employment and training is commendable, and is a quality that employers look for."
    p2 = "Your CareerGate Assessment results:"
    name = str(apprentice_vector[0]) + " " + str(apprentice_vector[1])
    p3 = "Assessment Subject: CareerGate 3 – General Aptitude"
    p4 = "Percentage: " + str(apprentice_vector[2])
    p5 = "Comment: " + comment
    p6 = "Date Taken: " + str(apprentice_vector[3])
    p7 = "Elapsed Time: " + str(apprentice_vector[4]) + " Minutes"

    # ADD TEXT
    story.append(Spacer(1, 30))
    H1 = Paragraph(h1, style=stylesheet["Heading_CENTER"])
    story.append(H1)

    story.append(Spacer(1, 50))
    P1 = Paragraph(p1, style=stylesheet['Text_CENTER'])
    story.append(P1)

    story.append(Spacer(1, 100))
    P2 = Paragraph(p2, style=stylesheet["Result_Heading_CENTER"])
    story.append(P2)

    story.append(Spacer(1, 10))
    NAME = Paragraph(name, style=stylesheet["Name_CENTER"])
    story.append(NAME)

    story.append(Spacer(1, 60))
    P3 = Paragraph(p3, style=stylesheet['Text_CENTER_BOLD'])
    story.append(P3)

    story.append(Spacer(1, 20))
    P4 = Paragraph(p4, style=stylesheet['Text_CENTER'])
    story.append(P4)

    story.append(Spacer(1, 10))
    P5 = Paragraph(p5, style=stylesheet['Text_CENTER'])
    story.append(P5)

    story.append(Spacer(1, 10))
    P6 = Paragraph(p6, style=stylesheet['Text_CENTER'])
    story.append(P6)

    story.append(Spacer(1, 10))
    P7 = Paragraph(p7, style=stylesheet['Text_CENTER'])
    story.append(P7)

    bottom_image = "Files\Gray_Underline.JPG"
    im = Image(bottom_image, 21 * cm, 1.2 * cm)
    im.__setattr__("_offs_x", 0)
    im.__setattr__("_offs_y", -210)
    story.append(im)

    pdf.build(story)

def generate_error_summary(error_rows, sheet):
    with open('error_summary.txt', 'w+') as f:
        f.write("ERRORS : \n\n")
        for i in range(0, len(error_rows)):
            if error_rows[i] != None:
                f.write(str(sheet.cell(error_rows[i], 1).value) + " " + str(sheet.cell(error_rows[i], 2).value) + "'s Certificate\n")

# *** SEQUENCE ***

#Check for Existing Worksheet -> If none -> Generate New
if not check_for_worksheet():
    generate_new_data_sheet()

#Load in workbook object
workbook = load_workbook(excel_file)
current_sheet = workbook.active

#Setup Memory to Store Apprentice Data
apprentice_count = get_apprentice_count(current_sheet)
    #Declare 2D array to hold data
apprentice_array = [[0] * column_count] * apprentice_count
error_row_array = [None] * apprentice_count

#Pull Data From Excel Sheet and Generate PDF
for i in range(0, apprentice_count):
    apprentice_array[i] = get_data(current_sheet, start_row + i)
    generate_pdf(apprentice_array[i])

if(len(error_row_array) > 0):
    generate_error_summary(error_row_array, current_sheet)









